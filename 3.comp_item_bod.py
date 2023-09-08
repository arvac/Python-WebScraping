import openpyxl
from logging import error
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
import time

username = "carmijo"
password = "Turion.3"
codigo = "p00095"
 
driver = webdriver.Chrome(executable_path="C:\chromedriver.exe" )
    # head to github login page
#driver.get("https://py3.favoritafruit.com:8185/jde/E1Menu.maf")
driver.get("https://jde.favoritafruit.com:8092/jde/E1Menu.maf?selectJPD920=SOP_ANLPD&envRadioGroup=&jdeowpBackButtonProtect=PROTECTED")

driver.implicitly_wait(0.5)
# find username/email field and send the username itself to the input field
#driver.find_element_by_id("login_field").send_keys(username)
driver.find_element("name", "User").send_keys(username)

# find password input field and insert password as well
#driver.find_element_by_id("password").send_keys(password)
driver.find_element("name", "Password").send_keys(password)
driver.find_element("name", "Password").send_keys(Keys.ENTER)
time.sleep(1)
   # click login button
driver.find_element("name" , "SUBMIT_BUTTON").click()

def proceso():
 workbook = openpyxl.load_workbook(r'C:\xampp\htdocs\py\comp\vacio.xlsx')
 sheet = workbook.active
 for row in sheet.iter_rows(min_row=2, values_only=True):
    
    time.sleep(2)
    item = row[0]
    bodega = row[1]

    time.sleep(1)
    driver.find_element("id", "drop_mainmenu").click()
    time.sleep(1)
    driver.find_element("id", "TE_FAST_PATH_BOX").send_keys("p4210")
    time.sleep(1)
    driver.find_element("id" , "fastPathButton").click()
    time.sleep(1)
    driver.switch_to.frame(8)

    for row in sheet.iter_rows(min_row=2, values_only=True):   
        time.sleep(1)
        item = row[0]
        bodega = row[1]
        time.sleep(1)
        #driver.switch_to.frame(8)

        driver.find_element(By.ID, "C0_13").click()
        driver.find_element(By.ID, "C0_13").clear()
        driver.find_element(By.ID, "C0_7").click()
        driver.find_element(By.ID, "C0_7").send_keys(bodega)
        time.sleep(1)
        driver.find_element(By.ID, "C0_31").click()
        driver.find_element(By.ID, "C0_31").send_keys(item)
        time.sleep(1)
        wait = WebDriverWait(driver, 10)
        select_element = wait.until(EC.presence_of_element_located((By.ID, "gtab0_1")))

# Utilizar el elemento Select para interactuar con el elemento select
        select = Select(select_element)

# Seleccionar la opción "Custom" por su valor
        select.select_by_value("a64851b0b1764fe1a71b1172f3a2117d")
        #select.select_by_value("a64851b0b1764fe1a71b1172f3a2117d")
        

        time.sleep(2)
        driver.find_element(By.NAME, "qbe0_1.1").clear()
        driver.find_element(By.NAME, "qbe0_1.1").send_keys("<580")
        driver.find_element(By.NAME, "qbe0_1.2").clear()
        driver.find_element(By.NAME, "qbe0_1.2").send_keys("!=S1")
        driver.find_element(By.NAME, "qbe0_1.3").clear()
        #driver.find_element(By.NAME, "qbe0_1.3").send_keys(">=15/07/23")
        driver.find_element(By.ID, "hc_Find").click()   
        time.sleep(3)


        try:
          estado = driver.find_element(By.XPATH,"//*[contains(text(), '580')]")
          
          driver.find_element(By.NAME, "hc_Close").click()
          driver.switch_to.default_content()
          time.sleep(2)
          driver.find_element(By.ID, "drop_mainmenu").click()
          time.sleep(1)
          driver.find_element(By.ID, "TE_FAST_PATH_BOX").send_keys("p4210")
          time.sleep(1)
          driver.find_element(By.ID, "fastPathButton").click()
                       
          driver.switch_to.frame(8) 
          #valor = estado.text
          # Buscar elementos que contengan los textos '560' o '580'
          #elementos = driver.find_elements(By.XPATH, "//*[contains(text(), '580')]")
          # Verificar si se encontraron elementos
          '''if len(elementos) <=580:
    # Se encontraron elementos que contienen '580' o valores menores
    # Capturar el valor '580' si existe

               print("Valor 580 encontrado:", elementos)
               driver.find_element(By.NAME, "hc_Close").click()
               driver.switch_to.default_content()
               time.sleep(2)
               driver.find_element(By.ID, "drop_mainmenu").click()
               time.sleep(1)
               driver.find_element(By.ID, "TE_FAST_PATH_BOX").send_keys("p4210")
               time.sleep(1)
               driver.find_element(By.ID, "fastPathButton").click()
                       
               driver.switch_to.frame(8) 
              
          else:
             print('PRINT')
          '''   
    # No se encontraron elementos que contengan '580' o valores menores
    # Ejecutar código adicional
     
                              
          
        except NoSuchElementException:
        #estado = driver.find_element(By.XPATH,"//table[@id='jdeGridData0_1.0']//td[contains(text(), '560')]")
        
          driver.find_element(By.NAME, "hc_Close").click()
          driver.switch_to.default_content()
          time.sleep(2)
          driver.find_element("id", "drop_mainmenu").click()
          time.sleep(1)
          driver.find_element("id", "TE_FAST_PATH_BOX").send_keys("bv")
          time.sleep(1)
          driver.find_element("id" , "fastPathButton").click()
          time.sleep(2)
          driver.switch_to.frame(8)
          driver.find_element(By.NAME, "0_11").send_keys("R5541078A")
          driver.find_element(By.ID, "hc_Find").click()
          driver.find_element(By.ID, "selectAll0_1").click()
          driver.find_element(By.ID, "hc_Select").click()
          time.sleep(3)
          driver.find_element(By.ID, "divC0_30").click()
          time.sleep(1)
          driver.find_element(By.ID, "PO8T0").click()
          driver.find_element(By.ID, "PO8T0").clear()
          driver.find_element(By.ID, "PO8T0").send_keys(item)
          driver.find_element(By.ID, "PO2T0").click()
          driver.find_element(By.ID, "PO2T0").send_keys(bodega)
          driver.find_element(By.ID, "PO3T0").click()
          driver.find_element(By.ID, "PO3T0").clear()
    #para campo vació y completos
          driver.find_element(By.ID, "PO3T0").send_keys(" ")
    #driver.find_element(By.ID, "PO3T0").send_keys(ubicacion)
          driver.find_element(By.ID, "PO4T0").click()
          driver.find_element(By.ID, "PO4T0").clear()
          driver.find_element(By.ID, "PO4T0").send_keys("")
    #driver.find_element(By.ID, "PO4T0").send_keys(lote)driver.find_element(By.ID, "PO7T0").click()
          driver.find_element(By.ID, "PO7T0").clear()
          driver.find_element(By.ID, "PO7T0").send_keys("A")
          driver.find_element(By.ID, "hc_Select").click()
          time.sleep(2)
          driver.find_element(By.ID, "hc_OK").click()
          time.sleep(2)
      
           #cierra
          driver.find_element(By.NAME, "hc_Close").click()
          driver.switch_to.default_content()
          time.sleep(2)
      
          driver.find_element("id", "drop_mainmenu").click()
          time.sleep(1)
          driver.find_element("id", "TE_FAST_PATH_BOX").send_keys("p4210")
          time.sleep(1)
          driver.find_element("id" , "fastPathButton").click()
          driver.switch_to.frame(8)
        '''  if "580" == valor  :
        # Ejecutar otro código si la comparación es verdadera
        # Aquí puedes agregar el código que deseas ejecutar cuando el texto es igual a "Texto1"
         driver.find_element(By.NAME, "hc_Close").click()
         driver.switch_to.default_content()
         time.sleep(2)
         driver.find_element("id", "drop_mainmenu").click()
         time.sleep(1)
         driver.find_element("id", "TE_FAST_PATH_BOX").send_keys("p4210")
         time.sleep(1)
         driver.find_element("id" , "fastPathButton").click()
         driver.switch_to.frame(8)
        else:
        '''   

proceso()
# Encontrar los elementos <span> por su atributo id
    #span_elements = driver.find_elements(By.CSS_SELECTOR, 'span[id="ActiveItemText"]')

# Iterar sobre los elementos <span> encontrados
'''for span_element in span_elements:
    # Obtener el texto contenido en el elemento <span>
    text = span_element.text
    
    # Comparar el texto con "Texto1"
    if text == "Texto1":
        # Ejecutar otro código si la comparación es verdadera
        # Aquí puedes agregar el código que deseas ejecutar cuando el texto es igual a "Texto1"
         driver.switch_to.frame(6)
         driver.find_element("id", "drop_mainmenu").click()
         time.sleep(1)
         driver.find_element("id", "TE_FAST_PATH_BOX").send_keys("p4210")
         time.sleep(1)
         driver.find_element("id" , "fastPathButton").click()
         time.sleep(1)
    else:
         
         driver.find_element("id", "drop_mainmenu").click()
         time.sleep(1)
         driver.find_element("id", "TE_FAST_PATH_BOX").send_keys("p00095")
         time.sleep(1)
         driver.find_element("id" , "fastPathButton").click()
'''
#driver.find_element(By.ID, "qbe0_1").click()
# driver.find_element(By.ID, "BUTTONCANCEL").click()
#driver.close()
